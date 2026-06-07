#!/usr/bin/env python3
"""
_tail_inject.py — Container-side data mutation script.

This script is COPY'd into the Docker container and executed during build.
It mutates input data files to introduce tail variants.

Supports both CSV and XLSX formats automatically.

Usage (in Dockerfile):
    COPY _tail_inject.py /tmp/_tail_inject.py
    RUN python3 /tmp/_tail_inject.py <variant> <filepath> [params...]
"""

import csv
import os
import random
import sys

# Fixed seed for reproducibility across builds
random.seed(42)


# ═══════════════════════════════════════════════════════
# File format helpers
# ═══════════════════════════════════════════════════════

def _is_xlsx(filepath):
    return filepath.lower().endswith(".xlsx")


def _read_tabular(filepath):
    """Read a CSV or XLSX file into (header, data_rows) where each row is a list of values."""
    if _is_xlsx(filepath):
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return [], []
        header = list(all_rows[0])
        data = [list(row) for row in all_rows[1:]]
        return header, data
    else:
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return [], []
        return rows[0], rows[1:]


def _write_tabular(filepath, header, data):
    """Write (header, data_rows) back to CSV or XLSX."""
    if _is_xlsx(filepath):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        for row in data:
            ws.append(row)
        wb.save(filepath)
    else:
        with open(filepath, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            writer.writerows(data)


# ═══════════════════════════════════════════════════════
# A1: BOM Injection
# ═══════════════════════════════════════════════════════
def inject_bom(filepath):
    """Inject UTF-8 BOM at the start of a text file."""
    if _is_xlsx(filepath):
        # For XLSX: export to CSV with BOM alongside the xlsx
        # The agent/skill may read this CSV instead of the xlsx
        csv_path = filepath.rsplit(".", 1)[0] + ".csv"
        header, data = _read_tabular(filepath)
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:  # utf-8-sig writes BOM
            writer = csv.writer(f)
            writer.writerow(header)
            writer.writerows(data)
        print(f"[TailSkills] BOM: exported {filepath} → {csv_path} with BOM")
        return

    with open(filepath, "rb") as f:
        content = f.read()
    if not content.startswith(b"\xef\xbb\xbf"):
        with open(filepath, "wb") as f:
            f.write(b"\xef\xbb\xbf" + content)
    print(f"[TailSkills] BOM injected → {filepath}")


# ═══════════════════════════════════════════════════════
# A2: Mixed Encoding (non-critical text fields only)
# ═══════════════════════════════════════════════════════
def inject_mixed_encoding(filepath, text_col_idx):
    """Replace a few words in a text column with Latin-1 encoded equivalents."""
    text_col_idx = int(text_col_idx)
    replacements = {
        "Region": "Région", "South": "Söuth", "North": "Nörth",
        "West": "Wëst", "New": "Nëw", "Lake": "Laké",
    }

    if _is_xlsx(filepath):
        # For XLSX: modify cell values in-place (stays valid Unicode in xlsx)
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=2):
            if text_col_idx < len(row) and count < 5:
                val = str(row[text_col_idx].value or "")
                for eng, lat in replacements.items():
                    if eng in val:
                        row[text_col_idx].value = val.replace(eng, lat, 1)
                        count += 1
                        break
        wb.save(filepath)
        print(f"[TailSkills] Mixed encoding (xlsx): {count} cells modified → {filepath}")
        return

    with open(filepath, "rb") as f:
        lines = f.readlines()
    count = 0
    for i in range(1, len(lines)):
        try:
            line_str = lines[i].decode("utf-8")
        except UnicodeDecodeError:
            continue
        for eng, lat in replacements.items():
            if eng in line_str and count < 5:
                line_str = line_str.replace(eng, lat, 1)
                lines[i] = line_str.encode("latin-1", errors="replace")
                count += 1
                break
    with open(filepath, "wb") as f:
        f.writelines(lines)
    print(f"[TailSkills] Mixed encoding: {count} lines modified → {filepath}")


# ═══════════════════════════════════════════════════════
# A3: Zero-Width Character Injection
# ═══════════════════════════════════════════════════════
def inject_zero_width(filepath, col_idx, ratio=0.03):
    """Insert zero-width characters into numeric field values."""
    col_idx = int(col_idx)
    ratio = float(ratio)
    zw_chars = ["\u200b", "\u200c", "\u200d"]

    header, data = _read_tabular(filepath)
    n_inject = max(1, int(len(data) * ratio))
    targets = random.sample(range(len(data)), min(n_inject, len(data)))

    for idx in targets:
        if col_idx < len(data[idx]):
            val = str(data[idx][col_idx] or "")
            if val and val.replace(".", "").replace("-", "").isdigit():
                mid = len(val) // 2
                data[idx][col_idx] = val[:mid] + random.choice(zw_chars) + val[mid:]

    _write_tabular(filepath, header, data)
    print(f"[TailSkills] Zero-width: {n_inject} values modified → {filepath}")


# ═══════════════════════════════════════════════════════
# C1: NaN/Inf Poisoning
# ═══════════════════════════════════════════════════════
def inject_nan_inf(filepath, col_idx, nan_ratio=0.02, inf_ratio=0.005):
    """Inject NaN and Inf values into a numeric column."""
    col_idx = int(col_idx)
    nan_ratio = float(nan_ratio)
    inf_ratio = float(inf_ratio)

    header, data = _read_tabular(filepath)
    n = len(data)

    nan_count = max(1, int(n * nan_ratio))
    inf_count = max(1, int(n * inf_ratio))

    nan_targets = set(random.sample(range(n), nan_count))
    inf_targets = set(random.sample(range(n), inf_count)) - nan_targets

    for idx in nan_targets:
        if col_idx < len(data[idx]):
            data[idx][col_idx] = None  # None → openpyxl writes empty cell → NaN
    for idx in inf_targets:
        if col_idx < len(data[idx]):
            data[idx][col_idx] = "inf"

    _write_tabular(filepath, header, data)
    print(f"[TailSkills] NaN/Inf: {len(nan_targets)} NaN + {len(inf_targets)} Inf → {filepath}")


# ═══════════════════════════════════════════════════════
# C3: Duplicate Key Injection
# ═══════════════════════════════════════════════════════
def inject_duplicate_keys(filepath, key_col_idx, n_dupes=3):
    """Duplicate a few rows keeping the same key but slightly different values."""
    key_col_idx = int(key_col_idx)
    n_dupes = int(n_dupes)

    header, data = _read_tabular(filepath)
    dupe_indices = random.sample(range(len(data)), min(n_dupes, len(data)))

    dupes = []
    for idx in dupe_indices:
        dupe = list(data[idx])
        for col in range(len(dupe)):
            if col != key_col_idx:
                try:
                    val = float(dupe[col])
                    dupe[col] = round(val * 1.05, 2)
                except (ValueError, TypeError):
                    pass
        dupes.append(dupe)

    data.extend(dupes)
    _write_tabular(filepath, header, data)
    print(f"[TailSkills] Duplicate keys: {len(dupes)} rows added → {filepath}")


# ═══════════════════════════════════════════════════════
# C4: Extreme Values
# ═══════════════════════════════════════════════════════
def inject_extremes(filepath, col_idx, n=3):
    """Inject zero, negative, and boundary values into a numeric column."""
    col_idx = int(col_idx)
    n = int(n)
    extremes = [0, -1, -999]

    header, data = _read_tabular(filepath)
    targets = random.sample(range(len(data)), min(n, len(data)))

    for i, idx in enumerate(targets):
        if col_idx < len(data[idx]):
            data[idx][col_idx] = extremes[i % len(extremes)]

    _write_tabular(filepath, header, data)
    print(f"[TailSkills] Extreme values: {n} injected → {filepath}")


# ═══════════════════════════════════════════════════════
# C5: Type Confusion
# ═══════════════════════════════════════════════════════
def inject_type_confusion(filepath, col_idx, n=5):
    """Replace numeric values with non-numeric markers like 'N/A', '-', etc."""
    col_idx = int(col_idx)
    n = int(n)
    markers = ["N/A", "n/a", "-", "#REF!", "TBD"]

    header, data = _read_tabular(filepath)
    targets = random.sample(range(len(data)), min(n, len(data)))

    for i, idx in enumerate(targets):
        if col_idx < len(data[idx]):
            data[idx][col_idx] = markers[i % len(markers)]

    _write_tabular(filepath, header, data)
    print(f"[TailSkills] Type confusion: {n} markers injected → {filepath}")


# ═══════════════════════════════════════════════════════
# A4: RFC 4180 Quoted Comma Injection
# ═══════════════════════════════════════════════════════
def inject_quoted_comma(filepath, text_col_idx, n=5):
    """Inject commas inside text field values (properly quoted per RFC 4180).

    The resulting CSV is fully standards-compliant — csv.reader and pd.read_csv
    handle it correctly. Only naive `line.split(',')` parsing will break.
    """
    text_col_idx = int(text_col_idx)
    n = int(n)

    suffixes = [
        ", supplementary info",
        ", revised",
        ", part A, part B",
        ", (see note 1, 2)",
        ", extra detail, updated",
    ]

    header, data = _read_tabular(filepath)
    targets = random.sample(range(len(data)), min(n, len(data)))

    count = 0
    for i, idx in enumerate(targets):
        if text_col_idx < len(data[idx]):
            val = str(data[idx][text_col_idx] or "")
            if val and val not in ("", "None", "nan"):
                data[idx][text_col_idx] = val + suffixes[i % len(suffixes)]
                count += 1

    # _write_tabular uses csv.writer which auto-quotes fields containing commas
    _write_tabular(filepath, header, data)
    print(f"[TailSkills] Quoted comma: {count} fields modified → {filepath}")


# ═══════════════════════════════════════════════════════
# A5: Date Format Diversity
# ═══════════════════════════════════════════════════════
def inject_date_diversity(filepath, date_col_idx, n=5):
    """Convert some YYYY-MM-DD dates to DD/MM/YYYY format.

    Only dates where day > 12 are converted, ensuring the result is
    unambiguous — both dayfirst=True and dayfirst=False will parse
    correctly, but a rigid `format='%Y-%m-%d'` will crash.
    """
    date_col_idx = int(date_col_idx)
    n = int(n)

    header, data = _read_tabular(filepath)

    # Collect candidate rows where date is YYYY-MM-DD and day > 12
    candidates = []
    for i, row in enumerate(data):
        if date_col_idx < len(row):
            val = str(row[date_col_idx] or "")
            parts = val.split("-")
            if len(parts) == 3:
                try:
                    year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
                    if day > 12 and 1 <= month <= 12:
                        candidates.append(i)
                except (ValueError, IndexError):
                    pass

    if not candidates:
        print(f"[TailSkills] Date diversity: no eligible dates found in {filepath}")
        return

    targets = random.sample(candidates, min(n, len(candidates)))
    count = 0
    for idx in targets:
        val = str(data[idx][date_col_idx])
        parts = val.split("-")
        # Convert YYYY-MM-DD → DD/MM/YYYY
        data[idx][date_col_idx] = f"{parts[2]}/{parts[1]}/{parts[0]}"
        count += 1

    _write_tabular(filepath, header, data)
    print(f"[TailSkills] Date diversity: {count} dates reformatted → {filepath}")


# ═══════════════════════════════════════════════════════
# Dispatcher
# ═══════════════════════════════════════════════════════
VARIANTS = {
    "bom": lambda args: inject_bom(args[0]),
    "mixed_encoding": lambda args: inject_mixed_encoding(args[0], args[1]),
    "zero_width": lambda args: inject_zero_width(args[0], args[1], *args[2:]),
    "nan_inf": lambda args: inject_nan_inf(args[0], args[1], *args[2:]),
    "duplicate_keys": lambda args: inject_duplicate_keys(args[0], args[1], *args[2:]),
    "extremes": lambda args: inject_extremes(args[0], args[1], *args[2:]),
    "type_confusion": lambda args: inject_type_confusion(args[0], args[1], *args[2:]),
    "quoted_comma": lambda args: inject_quoted_comma(args[0], args[1], *args[2:]),
    "date_diversity": lambda args: inject_date_diversity(args[0], args[1], *args[2:]),
}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <variant> <filepath> [params...]", file=sys.stderr)
        print(f"Available variants: {', '.join(VARIANTS.keys())}", file=sys.stderr)
        sys.exit(1)

    variant_name = sys.argv[1]
    rest = sys.argv[2:]

    if variant_name in VARIANTS:
        VARIANTS[variant_name](rest)
    else:
        print(f"[TailSkills] Unknown variant: {variant_name}", file=sys.stderr)
        print(f"[TailSkills] Available: {', '.join(VARIANTS.keys())}", file=sys.stderr)
        sys.exit(1)
