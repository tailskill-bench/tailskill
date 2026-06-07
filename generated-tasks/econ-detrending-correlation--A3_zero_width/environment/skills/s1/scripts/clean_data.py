#!/usr/bin/env python3
"""Strip zero-width Unicode characters from XLSX/CSV files."""
import sys
import os
import unicodedata


def strip_zw_xlsx(filepath):
    from openpyxl import load_workbook
    wb = load_workbook(filepath)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cleaned = ''.join(c for c in cell.value if unicodedata.category(c) not in ('Cf', 'Cc'))
                    if cleaned != cell.value:
                        try:
                            cell.value = float(cleaned)
                        except ValueError:
                            cell.value = cleaned
    wb.save(filepath)
    print(f"Zero-width chars stripped from {filepath}")


def strip_zw_csv(filepath):
    import csv
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))
    changed = False
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            if isinstance(val, str):
                cleaned = ''.join(c for c in val if unicodedata.category(c) not in ('Cf', 'Cc'))
                if cleaned != val:
                    rows[i][j] = cleaned
                    changed = True
    if changed:
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            csv.writer(f).writerows(rows)
        print(f"Zero-width chars stripped from {filepath}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: clean_data.py <file>")
        sys.exit(1)
    target = sys.argv[1]
    if target.endswith('.xlsx'):
        strip_zw_xlsx(target)
    elif target.endswith('.csv'):
        strip_zw_csv(target)
    else:
        print(f"Unsupported file type: {target}")
