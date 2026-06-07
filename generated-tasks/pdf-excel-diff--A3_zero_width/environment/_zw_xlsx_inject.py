#!/usr/bin/env python3
"""Inject zero-width Unicode characters into XLSX string cells."""
import random
import sys
from openpyxl import load_workbook

ZW_CHARS = ['\u200b', '\u200c', '\u200d', '\ufeff']


def inject_zero_width_xlsx(filepath, ratio=0.03):
    random.seed(42)
    wb = load_workbook(filepath)
    ws = wb.active
    n_injected = 0

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str) and len(cell.value) > 0:
                if random.random() < ratio:
                    pos = random.randint(0, len(cell.value))
                    cell.value = cell.value[:pos] + random.choice(ZW_CHARS) + cell.value[pos:]
                    n_injected += 1

    wb.save(filepath)
    print(f"[TailSkills] A3_zero_width: {n_injected} zero-width chars injected → {filepath}")


if __name__ == "__main__":
    if len(sys.argv) >= 2:
        inject_zero_width_xlsx(sys.argv[1])
    else:
        print("Usage: _zw_xlsx_inject.py <xlsx_file>")
