#!/usr/bin/env python3
"""Strip zero-width Unicode characters from XLSX string cells."""
import sys
import unicodedata
from openpyxl import load_workbook

ZW_CHARS = {'\u200b', '\u200c', '\u200d', '\ufeff'}


def strip_zero_width_xlsx(filepath):
    wb = load_workbook(filepath)
    n_fixed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str):
                    cleaned = ''.join(
                        c for c in cell.value
                        if c not in ZW_CHARS and unicodedata.category(c) != 'Cf'
                    )
                    if cleaned != cell.value:
                        cell.value = cleaned
                        n_fixed += 1
    wb.save(filepath)
    print(f"Stripped zero-width chars from {n_fixed} cells in {filepath}")


if __name__ == "__main__":
    if len(sys.argv) >= 2:
        strip_zero_width_xlsx(sys.argv[1])
    else:
        print("Usage: clean_data.py <xlsx_file>")
