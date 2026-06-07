#!/usr/bin/env python3
"""Inject NaN into a numeric column of XLSX by clearing cells (sets to None)."""
import random
import sys
from openpyxl import load_workbook


def inject_nan_xlsx(filepath, col_idx, ratio=0.02):
    random.seed(42)
    col_idx = int(col_idx)
    wb = load_workbook(filepath)
    ws = wb.active
    n_nan = 0

    rows = list(ws.iter_rows(min_row=2))
    for row in rows:
        cell = row[col_idx]
        if cell.value is not None and isinstance(cell.value, (int, float)):
            r = random.random()
            if r < ratio * 1.25:
                cell.value = None
                n_nan += 1

    wb.save(filepath)
    print(f"[TailSkills] NaN: {n_nan} cells cleared (set to None) → {filepath} (col {col_idx})")


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        inject_nan_xlsx(sys.argv[1], sys.argv[2])
    else:
        print("Usage: _nan_xlsx_inject.py <xlsx_file> <col_idx>")
