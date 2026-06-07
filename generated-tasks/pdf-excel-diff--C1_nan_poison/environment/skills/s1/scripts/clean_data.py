#!/usr/bin/env python3
"""Clean NaN/Inf/None values from XLSX numeric columns using median fillna."""
import math
import sys
from openpyxl import load_workbook


def clean_nan_xlsx(filepath, col_idx):
    col_idx = int(col_idx)
    wb = load_workbook(filepath)
    ws = wb.active

    # Compute median of valid numeric values
    valid = []
    for row in ws.iter_rows(min_row=2):
        cell = row[col_idx]
        if cell.value is not None and isinstance(cell.value, (int, float)):
            v = float(cell.value)
            if not (math.isnan(v) or math.isinf(v)):
                valid.append(v)

    if not valid:
        print(f"No valid values in column {col_idx}, skipping")
        return

    valid.sort()
    median = valid[len(valid) // 2]

    # Replace None, NaN, Inf with median
    n_fixed = 0
    for row in ws.iter_rows(min_row=2):
        cell = row[col_idx]
        if cell.value is None:
            cell.value = median
            n_fixed += 1
        elif isinstance(cell.value, float):
            if math.isnan(cell.value) or math.isinf(cell.value):
                cell.value = median
                n_fixed += 1

    wb.save(filepath)
    print(f"Fixed {n_fixed} NaN/Inf/None values in column {col_idx} with median={median}")


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        clean_nan_xlsx(sys.argv[1], sys.argv[2])
    else:
        print("Usage: clean_data.py <xlsx_file> <col_idx>")
