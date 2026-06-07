#!/usr/bin/env python3
"""Inject extreme values (0, -1, -999, 999999) into a numeric column of XLSX."""
import random
import sys
from openpyxl import load_workbook

EXTREME_VALUES = [-999, -1, 999999]


def inject_extremes_xlsx(filepath, col_idx, ratio=0.025):
    random.seed(42)
    col_idx = int(col_idx)
    wb = load_workbook(filepath)
    ws = wb.active
    n_extreme = 0

    rows = list(ws.iter_rows(min_row=2))
    for row in rows:
        cell = row[col_idx]
        if cell.value is not None and isinstance(cell.value, (int, float)):
            if random.random() < ratio * 1.25:
                cell.value = random.choice(EXTREME_VALUES)
                n_extreme += 1

    wb.save(filepath)
    print(f"[TailSkills] Extremes: {n_extreme} cells injected → {filepath} (col {col_idx})")


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        inject_extremes_xlsx(sys.argv[1], sys.argv[2])
    else:
        print("Usage: _extremes_xlsx_inject.py <xlsx_file> <col_idx>")
