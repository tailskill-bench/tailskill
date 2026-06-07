#!/usr/bin/env python3
"""Clean extreme values from XLSX numeric columns by replacing with None (so comparison skips them)."""
import sys
from openpyxl import load_workbook


def clean_extremes_xlsx(filepath, col_idx, low=0, high=100):
    col_idx = int(col_idx)
    wb = load_workbook(filepath)
    ws = wb.active

    n_fixed = 0
    for row in ws.iter_rows(min_row=2):
        cell = row[col_idx]
        if cell.value is not None and isinstance(cell.value, (int, float)):
            v = float(cell.value)
            if v < low or v > high:
                cell.value = None
                n_fixed += 1

    wb.save(filepath)
    print(f"Fixed {n_fixed} extreme values in column {col_idx} (set to None, outside [{low}, {high}])")


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        clean_extremes_xlsx(sys.argv[1], sys.argv[2])
    else:
        print("Usage: clean_data.py <xlsx_file> <col_idx>")
