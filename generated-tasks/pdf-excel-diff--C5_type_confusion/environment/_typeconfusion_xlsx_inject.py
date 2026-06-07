#!/usr/bin/env python3
"""Inject type confusion into a numeric column of XLSX by replacing with text markers."""
import random
import sys
from openpyxl import load_workbook

TEXT_MARKERS = ["N/A", "-", "TBD", "null", "none"]


def inject_typeconfusion_xlsx(filepath, col_idx, ratio=0.025):
    random.seed(42)
    col_idx = int(col_idx)
    wb = load_workbook(filepath)
    ws = wb.active
    n_injected = 0

    rows = list(ws.iter_rows(min_row=2))
    for row in rows:
        cell = row[col_idx]
        if cell.value is not None and isinstance(cell.value, (int, float)):
            if random.random() < ratio * 1.25:
                cell.value = random.choice(TEXT_MARKERS)
                n_injected += 1

    wb.save(filepath)
    print(f"[TailSkills] TypeConfusion: {n_injected} cells injected → {filepath} (col {col_idx})")


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        inject_typeconfusion_xlsx(sys.argv[1], sys.argv[2])
    else:
        print("Usage: _typeconfusion_xlsx_inject.py <xlsx_file> <col_idx>")
