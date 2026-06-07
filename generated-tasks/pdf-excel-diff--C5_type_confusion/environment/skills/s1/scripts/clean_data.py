#!/usr/bin/env python3
"""Clean type confusion from XLSX numeric columns by replacing text markers with None."""
import sys
from openpyxl import load_workbook

TEXT_MARKERS = {"n/a", "-", "tbd", "null", "none", "", "nan", "na"}


def clean_typeconfusion_xlsx(filepath, col_idx):
    col_idx = int(col_idx)
    wb = load_workbook(filepath)
    ws = wb.active

    n_fixed = 0
    for row in ws.iter_rows(min_row=2):
        cell = row[col_idx]
        if cell.value is not None and isinstance(cell.value, str):
            if cell.value.strip().lower() in TEXT_MARKERS:
                cell.value = None
                n_fixed += 1

    wb.save(filepath)
    print(f"Fixed {n_fixed} type-confused values in column {col_idx} (set to None)")


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        clean_typeconfusion_xlsx(sys.argv[1], sys.argv[2])
    else:
        print("Usage: clean_data.py <xlsx_file> <col_idx>")
