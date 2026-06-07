import random, math
random.seed(42)
from openpyxl import load_workbook

wb = load_workbook('/root/data/openipf.xlsx')
ws = wb.active
col_idx = 8  # 0-indexed column 8 = column I (1-indexed 9)
n_rows = ws.max_row
n_inject = max(1, int(n_rows * 0.02))
targets = random.sample(range(2, n_rows + 1), min(n_inject, n_rows - 1))
for row_idx in targets:
    cell = ws.cell(row=row_idx, column=col_idx + 1)
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.value = float('nan')
wb.save('/root/data/openipf.xlsx')
print(f'[TailSkills] NaN: {len(targets)} cells modified in column {col_idx+1}')
