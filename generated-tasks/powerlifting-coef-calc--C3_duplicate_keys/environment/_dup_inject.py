import random
random.seed(42)
from openpyxl import load_workbook

wb = load_workbook('/root/data/openipf.xlsx')
ws = wb.active
n_rows = ws.max_row
n_dupes = 3
# Pick random source rows (skip header)
source_rows = random.sample(range(2, n_rows + 1), min(n_dupes, n_rows - 1))
for src_row in source_rows:
    new_row = n_rows + 1
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=src_row, column=col).value
        # Slightly alter numeric non-key columns
        if col > 1 and isinstance(val, (int, float)):
            val = round(val * 1.05, 2)
        ws.cell(row=new_row, column=col).value = val
    n_rows += 1
wb.save('/root/data/openipf.xlsx')
print(f'[TailSkills] Duplicate keys: {len(source_rows)} rows added')
