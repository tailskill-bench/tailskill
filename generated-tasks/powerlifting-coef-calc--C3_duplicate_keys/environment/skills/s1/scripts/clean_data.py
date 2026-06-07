#!/usr/bin/env python3
"""Drop duplicate rows by key column — proper error handling for duplicates."""
import sys
from openpyxl import load_workbook

target = sys.argv[1] if len(sys.argv) > 1 else "/root/data/openipf.xlsx"
key_col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 0

wb = load_workbook(target)
ws = wb["Data"] if "Data" in wb.sheetnames else wb.active

seen = set()
rows_to_delete = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    key = row[key_col_idx].value if key_col_idx < len(row) else None
    if key is not None and key in seen:
        rows_to_delete.append(row_idx)
    elif key is not None:
        seen.add(key)

for row_idx in reversed(rows_to_delete):
    ws.delete_rows(row_idx)

wb.save(target)
print(f"Dropped {len(rows_to_delete)} duplicate rows by key column {key_col_idx}")
