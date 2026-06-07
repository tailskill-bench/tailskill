import random
random.seed(42)
from openpyxl import load_workbook

wb = load_workbook('/root/data/openipf.xlsx')
ws = wb.active
col_idx = 11  # 1-indexed column 11 = column K (Best3SquatKg), 0-indexed 10
n_rows = ws.max_row
extremes = [0, -1, -999]
targets = random.sample(range(2, n_rows + 1), 3)
for i, row_idx in enumerate(targets):
    cell = ws.cell(row=row_idx, column=col_idx)
    if cell.value is not None:
        cell.value = extremes[i % len(extremes)]
wb.save('/root/data/openipf.xlsx')
print(f'[TailSkills] Extremes: 3 cells modified in column {col_idx} with values {extremes}')
