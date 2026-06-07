#!/usr/bin/env python3
"""Drop rows with extreme outlier values — proper error handling using MAD detection."""
import sys, math
from openpyxl import load_workbook

target = sys.argv[1] if len(sys.argv) > 1 else "/root/data/openipf.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 10

wb = load_workbook(target)
ws = wb["Data"] if "Data" in wb.sheetnames else wb.active

valid = []
for row in ws.iter_rows(min_row=2, values_only=False):
    cell = row[col_idx] if col_idx < len(row) else None
    if cell and isinstance(cell.value, (int, float)) and math.isfinite(cell.value):
        valid.append(cell.value)

if not valid:
    wb.save(target)
    print("No valid values found")
    sys.exit(0)

median = sorted(valid)[len(valid)//2]
mad = sorted(abs(v - median) for v in valid)[len(valid)//2]
threshold = max(5 * mad, 1.0)

rows_to_delete = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    cell = row[col_idx] if col_idx < len(row) else None
    if cell is None:
        continue
    val = cell.value
    if val is None or isinstance(val, str):
        rows_to_delete.append(row_idx)
    elif isinstance(val, (int, float)):
        if not math.isfinite(val) or abs(val - median) > threshold or val <= 0:
            rows_to_delete.append(row_idx)

for row_idx in reversed(rows_to_delete):
    ws.delete_rows(row_idx)

wb.save(target)
print(f"Dropped {len(rows_to_delete)} rows with extreme values (threshold={threshold:.2f})")
