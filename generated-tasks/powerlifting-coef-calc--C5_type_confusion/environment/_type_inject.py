import random
random.seed(42)
from openpyxl import load_workbook

wb = load_workbook('/root/data/openipf.xlsx')
ws = wb.active
col_idx = 9  # 1-indexed column 9 = column I (BodyweightKg), 0-indexed 8
n_rows = ws.max_row
markers = ["N/A", "n/a", "-", "TBD", "#REF!"]
targets = random.sample(range(2, n_rows + 1), min(5, n_rows - 1))
for i, row_idx in enumerate(targets):
    cell = ws.cell(row=row_idx, column=col_idx)
    if cell.value is not None:
        cell.value = markers[i % len(markers)]
wb.save('/root/data/openipf.xlsx')
print(f'[TailSkills] Type confusion: {len(targets)} cells modified in column {col_idx}')
