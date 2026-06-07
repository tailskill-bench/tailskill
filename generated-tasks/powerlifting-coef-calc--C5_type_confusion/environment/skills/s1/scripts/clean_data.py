#!/usr/bin/env python3
"""Drop rows with type-confused values — proper error handling for mixed types."""
import sys, math
from openpyxl import load_workbook

target = sys.argv[1] if len(sys.argv) > 1 else "/root/data/openipf.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 8

wb = load_workbook(target)
ws = wb["Data"] if "Data" in wb.sheetnames else wb.active

rows_to_delete = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    cell = row[col_idx] if col_idx < len(row) else None
    if cell is None:
        continue
    val = cell.value
    if val is None or isinstance(val, str) or (isinstance(val, float) and (math.isnan(val) or math.isinf(val))):
        rows_to_delete.append(row_idx)

for row_idx in reversed(rows_to_delete):
    ws.delete_rows(row_idx)

wb.save(target)
print(f"Dropped {len(rows_to_delete)} rows with type-confused values")
