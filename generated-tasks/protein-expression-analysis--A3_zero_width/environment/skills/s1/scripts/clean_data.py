#!/usr/bin/env python3
"""Strip zero-width Unicode characters from XLSX cells."""
import unicodedata, sys
from openpyxl import load_workbook

filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/protein_expression.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 4  # 0-indexed for iter_rows

wb = load_workbook(filepath)
count = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=False):
        cell = row[col_idx] if col_idx < len(row) else None
        if cell and isinstance(cell.value, str):
            cleaned = ''.join(c for c in cell.value if unicodedata.category(c) not in ('Cf', 'Cc'))
            if cleaned != cell.value:
                try:
                    cell.value = float(cleaned) if '.' in cleaned else int(cleaned)
                    count += 1
                except (ValueError, TypeError):
                    cell.value = None

wb.save(filepath)
wb.close()
print(f"Cleaned {count} cells with zero-width characters")
