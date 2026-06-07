#!/usr/bin/env python3
"""Inject extreme values into Data sheet numeric cells."""
import random, sys
from openpyxl import load_workbook

random.seed(42)
filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/protein_expression.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 5
extremes = [0, -1, -999]

wb = load_workbook(filepath)
ws = wb["Data"]

numeric_rows = [r for r in range(2, ws.max_row + 1)
                if isinstance(ws.cell(row=r, column=col_idx).value, (int, float))]

targets = random.sample(numeric_rows, min(3, len(numeric_rows)))
for i, row in enumerate(targets):
    ws.cell(row=row, column=col_idx).value = extremes[i % len(extremes)]

wb.save(filepath)
print(f"[TailSkills] Extremes: {len(targets)} cells modified in Data sheet col {col_idx}")
