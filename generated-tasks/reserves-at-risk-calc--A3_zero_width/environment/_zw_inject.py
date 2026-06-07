#!/usr/bin/env python3
"""Inject zero-width Unicode chars into XLSX numeric cells (preserves workbook structure)."""
import random, sys, unicodedata
from openpyxl import load_workbook

random.seed(42)
filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/data/test-rar.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 4  # 0-indexed
zw_chars = ['\u200b', '\u200c', '\u200d']

wb = load_workbook(filepath)
target_col = col_idx + 1  # 1-indexed

count = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    numeric_rows = []
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=target_col)
        if isinstance(cell.value, (int, float)):
            numeric_rows.append(r)

    if not numeric_rows:
        continue
    targets = random.sample(numeric_rows, min(max(1, len(numeric_rows) // 10), 5))
    for row in targets:
        cell = ws.cell(row=row, column=target_col)
        zw = random.choice(zw_chars)
        cell.value = zw + str(cell.value) + zw
        count += 1

wb.save(filepath)
print(f"[TailSkills] Zero-width: {count} cells modified across {len(wb.sheetnames)} sheets col {target_col}")
