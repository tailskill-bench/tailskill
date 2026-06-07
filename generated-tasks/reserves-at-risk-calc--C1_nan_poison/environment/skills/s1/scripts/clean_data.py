#!/usr/bin/env python3
"""Replace NaN/Inf/None values in specific data sheets with column-median."""
import sys, math
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/data/test-rar.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 4  # 0-indexed

# Only clean input data sheets (not Gold price or Answer which are written by solve)
TARGET_SHEETS = ["Value", "Volume", "Total Reserves"]

wb = load_workbook(filepath)
target_col = col_idx + 1  # 1-indexed

for sheet_name in wb.sheetnames:
    if sheet_name not in TARGET_SHEETS:
        continue
    ws = wb[sheet_name]
    # Compute column median from valid numeric values (skip header row 1)
    valid = []
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=target_col)
        if isinstance(cell, MergedCell):
            continue
        v = cell.value
        if v is not None and isinstance(v, (int, float)) and math.isfinite(v):
            valid.append(v)

    if not valid:
        continue
    col_median = sorted(valid)[len(valid) // 2]

    count = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=target_col)
        if isinstance(cell, MergedCell):
            continue
        val = cell.value
        if val is None or (isinstance(val, float) and not math.isfinite(val)) or isinstance(val, str):
            cell.value = col_median
            count += 1

    if count > 0:
        print(f"  {sheet_name}: replaced {count} values with median ({col_median})")

wb.save(filepath)
wb.close()
print(f"Cleaned NaN/Inf values in {filepath}")
