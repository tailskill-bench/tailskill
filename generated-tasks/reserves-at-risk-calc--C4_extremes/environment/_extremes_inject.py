#!/usr/bin/env python3
"""Inject extreme values into XLSX numeric cells (preserves workbook structure)."""
import random, sys
from openpyxl import load_workbook

random.seed(42)
filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/data/test-rar.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 4  # 0-indexed
extremes = [0, -1, -999]

wb = load_workbook(filepath)
target_col = col_idx + 1  # 1-indexed

count = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Skip header row 1
    numeric_rows = [r for r in range(2, ws.max_row + 1)
                    if isinstance(ws.cell(row=r, column=target_col).value, (int, float))]
    if not numeric_rows:
        continue
    targets = random.sample(numeric_rows, min(3, len(numeric_rows)))
    for i, row in enumerate(targets):
        cell = ws.cell(row=row, column=target_col)
        cell.value = extremes[i % len(extremes)]
        count += 1

wb.save(filepath)
print(f"[TailSkills] Extremes: {count} cells modified in col {target_col}")
