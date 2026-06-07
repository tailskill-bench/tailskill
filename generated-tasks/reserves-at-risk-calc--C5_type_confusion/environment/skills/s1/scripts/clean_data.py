#!/usr/bin/env python3
"""Convert string-typed numeric cells back to proper numeric types in data sheets."""
import sys
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/data/test-rar.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 4  # 0-indexed

TARGET_SHEETS = ["Value", "Volume", "Total Reserves"]

wb = load_workbook(filepath)
target_col = col_idx + 1  # 1-indexed

for sheet_name in wb.sheetnames:
    if sheet_name not in TARGET_SHEETS:
        continue
    ws = wb[sheet_name]

    count = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=target_col)
        if isinstance(cell, MergedCell):
            continue
        val = cell.value
        if isinstance(val, str):
            try:
                num = float(val)
                if num == int(num):
                    num = int(num)
                cell.value = num
                count += 1
            except ValueError:
                pass

    if count > 0:
        print(f"  {sheet_name}: fixed {count} type-confused cells")

wb.save(filepath)
wb.close()
print(f"Fixed type confusion in {filepath}")
