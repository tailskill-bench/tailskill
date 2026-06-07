#!/usr/bin/env python3
"""Data quality helper — removes invisible Unicode characters from numeric fields in CSV/XLSX files."""

import sys
import unicodedata
from pathlib import Path


def _is_xlsx(filepath):
    return filepath.lower().endswith(".xlsx")


def clean_zero_width(filepath, col_idx=None):
    """Remove zero-width and format Unicode characters from numeric fields."""
    if _is_xlsx(filepath):
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    cleaned = ''.join(
                        c for c in cell.value
                        if unicodedata.category(c) not in ('Cf', 'Cc')
                    )
                    if cleaned != cell.value:
                        cell.value = cleaned
                        count += 1
        wb.save(filepath)
        print(f"[clean_data] Zero-width: {count} cells cleaned in {filepath}")
    else:
        import csv
        with open(filepath, 'r', encoding='utf-8', newline='') as f:
            rows = list(csv.reader(f))
        count = 0
        for row in rows:
            for i, val in enumerate(row):
                cleaned = ''.join(
                    c for c in val
                    if unicodedata.category(c) not in ('Cf', 'Cc')
                )
                if cleaned != val:
                    row[i] = cleaned
                    count += 1
        with open(filepath, 'w', encoding='utf-8', newline='') as f:
            csv.writer(f).writerows(rows)
        print(f"[clean_data] Zero-width: {count} cells cleaned in {filepath}")


def main():
    if len(sys.argv) < 2:
        print("Usage: clean_data.py <file> [--col COL_IDX]")
        sys.exit(1)

    target = sys.argv[1]
    col_idx = None
    if '--col' in sys.argv:
        col_idx = int(sys.argv[sys.argv.index('--col') + 1])

    clean_zero_width(target, col_idx)


if __name__ == "__main__":
    main()
