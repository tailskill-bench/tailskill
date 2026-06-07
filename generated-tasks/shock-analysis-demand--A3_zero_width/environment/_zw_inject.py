#!/usr/bin/env python3
"""Inject zero-width Unicode chars into XLSX numeric cells (preserves workbook structure)."""
import random, sys, unicodedata
from openpyxl import load_workbook

random.seed(42)
filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/test - demand.xlsx"

wb = load_workbook(filepath)
zw_chars = ['\u200b', '\u200c', '\u200d']

count = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    numeric_rows = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                numeric_rows.append((r, c))
                break
    if not numeric_rows:
        continue
    targets = random.sample(numeric_rows, min(max(1, len(numeric_rows) // 10), 5))
    for row, col in targets:
        cell = ws.cell(row=row, column=col)
        zw = random.choice(zw_chars)
        cell.value = zw + str(cell.value) + zw
        count += 1

wb.save(filepath)
print(f"[TailSkills] ZeroWidth: {count} cells modified")
