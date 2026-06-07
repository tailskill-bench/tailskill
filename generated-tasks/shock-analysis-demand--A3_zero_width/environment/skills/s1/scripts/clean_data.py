#!/usr/bin/env python3
"""Strip zero-width Unicode characters from XLSX cells and restore numeric types."""
import sys, unicodedata
from openpyxl import load_workbook

filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/test - demand.xlsx"
wb = load_workbook(filepath)
count = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cleaned = ''.join(c for c in cell.value if unicodedata.category(c) not in ('Cf', 'Cc'))
                if cleaned != cell.value:
                    try:
                        cell.value = float(cleaned)
                        if cell.value == int(cell.value):
                            cell.value = int(cell.value)
                        count += 1
                    except ValueError:
                        cell.value = cleaned
                        count += 1
wb.save(filepath)
wb.close()
print(f"Cleaned {count} cells with zero-width characters")
