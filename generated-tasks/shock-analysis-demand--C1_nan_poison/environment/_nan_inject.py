#!/usr/bin/env python3
"""Inject NaN/Inf into XLSX numeric cells (preserves workbook structure)."""
import random, sys, math
from openpyxl import load_workbook

random.seed(42)
filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/test - demand.xlsx"

wb = load_workbook(filepath)

count = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    numeric_rows = [r for r in range(2, ws.max_row + 1)
                    if any(isinstance(ws.cell(row=r, column=c).value, (int, float))
                           for c in range(1, min(ws.max_column + 1, 10)))]
    if not numeric_rows:
        continue
    targets = random.sample(numeric_rows, min(3, len(numeric_rows)))
    for i, row in enumerate(targets):
        for c in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row=row, column=c)
            if isinstance(cell.value, (int, float)):
                if i < 2:
                    cell.value = None
                else:
                    cell.value = float('inf')
                count += 1
                break

wb.save(filepath)
print(f"[TailSkills] NaN/Inf: {count} cells modified")
