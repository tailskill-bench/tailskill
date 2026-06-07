#!/usr/bin/env python3
"""Replace NaN/Inf/None values in XLSX with column-median."""
import sys, math
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/test - demand.xlsx"
wb = load_workbook(filepath)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for col in range(1, ws.max_column + 1):
        valid = []
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if v is not None and isinstance(v, (int, float)) and math.isfinite(v):
                valid.append(v)
        if not valid:
            continue
        col_median = sorted(valid)[len(valid) // 2]

        count = 0
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if val is None or (isinstance(val, float) and not math.isfinite(val)) or isinstance(val, str):
                cell.value = col_median
                count += 1
        if count > 0:
            print(f"  {sheet_name} col {col}: replaced {count} values with median ({col_median})")

wb.save(filepath)
wb.close()
print(f"Cleaned NaN/Inf values in {filepath}")
