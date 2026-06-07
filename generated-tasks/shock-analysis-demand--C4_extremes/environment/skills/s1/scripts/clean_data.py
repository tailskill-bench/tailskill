#!/usr/bin/env python3
"""Replace extreme values in XLSX with column-median using MAD + non-positive detection."""
import sys, math
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/test - demand.xlsx"
wb = load_workbook(filepath)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for col in range(1, ws.max_column + 1):
        valid = []
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if v is not None and isinstance(v, (int, float)) and math.isfinite(v) and v > 0:
                valid.append(v)
        if not valid:
            continue
        median_col = sorted(valid)[len(valid) // 2]
        mad = sorted(abs(v - median_col) for v in valid)[len(valid) // 2]
        threshold = max(5 * mad, 1.0)

        count = 0
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            corrupted = False
            if val is None:
                corrupted = True
            elif isinstance(val, str):
                corrupted = True
            elif isinstance(val, (int, float)):
                if not math.isfinite(val) or val <= 0 or abs(val - median_col) > threshold:
                    corrupted = True
            if corrupted:
                cell.value = median_col
                count += 1
        if count > 0:
            print(f"  {sheet_name} col {col}: replaced {count} extreme values")

wb.save(filepath)
wb.close()
print(f"Cleaned extreme values in {filepath}")
