#!/usr/bin/env python3
"""Inject type confusion into XLSX numeric cells — convert numbers to strings."""
import random, sys
from openpyxl import load_workbook

random.seed(42)
filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/test - demand.xlsx"

wb = load_workbook(filepath)

count = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    numeric_rows = [r for r in range(2, ws.max_row + 1)
                    if any(isinstance(ws.cell(row=r, column=c).value, (int, float))
                           for c in range(1, min(ws.max_column + 1, 10)))]
    if not numeric_rows:
        continue
    targets = random.sample(numeric_rows, min(3, len(numeric_rows)))
    for row in targets:
        for c in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row=row, column=c)
            if isinstance(cell.value, (int, float)):
                cell.value = str(cell.value)
                count += 1
                break

wb.save(filepath)
print(f"[TailSkills] TypeConfusion: {count} cells converted to string")
