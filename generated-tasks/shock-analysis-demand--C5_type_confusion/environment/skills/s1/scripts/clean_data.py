#!/usr/bin/env python3
"""Convert string-typed numeric cells back to proper numeric types in XLSX."""
import sys
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/test - demand.xlsx"
wb = load_workbook(filepath)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    count = 0
    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if isinstance(val, str):
                try:
                    num = float(val)
                    if num == int(num):
                        num = int(num)
                    cell.value = num
                    count += 1
                except ValueError:
                    pass
    if count > 0:
        print(f"  {sheet_name}: fixed {count} type-confused cells")

wb.save(filepath)
wb.close()
print(f"Fixed type confusion in {filepath}")
