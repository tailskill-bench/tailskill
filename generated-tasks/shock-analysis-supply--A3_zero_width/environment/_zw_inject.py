"""A3_zero_width injector: Inject zero-width Unicode chars into numeric cells of test-supply.xlsx."""
import random, unicodedata
from openpyxl import load_workbook

ZW_CHARS = [unicodedata.lookup(n) for n in ['ZERO WIDTH SPACE', 'ZERO WIDTH JOINER', 'ZERO WIDTH NON-JOINER', 'WORD JOINER']]
random.seed(42)
wb = load_workbook('/root/test-supply.xlsx')

for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                zw = random.choice(ZW_CHARS)
                cell.value = f"{zw}{cell.value}{zw}"

wb.save('/root/test-supply.xlsx')
print("Injected zero-width chars into test-supply.xlsx")
