"""Strip zero-width Unicode characters from XLSX cells and restore numeric types."""
import unicodedata
from openpyxl import load_workbook

path = '/root/test-supply.xlsx'
wb = load_workbook(path)

for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str):
                cleaned = ''.join(c for c in cell.value if unicodedata.category(c) not in ('Cf', 'Cc'))
                try:
                    cell.value = float(cleaned)
                    if cell.value == int(cell.value):
                        cell.value = int(cell.value)
                except (ValueError, TypeError):
                    cell.value = cleaned

wb.save(path)
print(f"Cleaned zero-width chars in {path}")
