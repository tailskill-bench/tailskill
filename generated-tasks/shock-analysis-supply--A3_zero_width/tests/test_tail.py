import unicodedata, pytest
from openpyxl import load_workbook

class TestTailVariant:
    def test_no_zero_width_chars(self):
        wb = load_workbook('/root/test-supply.xlsx')
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        for c in cell.value:
                            assert unicodedata.category(c) != 'Cf', f"Zero-width char in {ws.title}!{cell.coordinate}"
