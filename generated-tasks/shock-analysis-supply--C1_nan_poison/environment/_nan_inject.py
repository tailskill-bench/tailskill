"""C1_nan_poison injector: Inject NaN/Inf/None into numeric cells of test-supply.xlsx, preserving sheet structure."""
import random, math
from openpyxl import load_workbook

random.seed(42)
wb = load_workbook('/root/test-supply.xlsx')
poisons = [float('nan'), float('inf'), float('-inf'), None]

for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and random.random() < 0.15:
                cell.value = random.choice(poisons)

wb.save('/root/test-supply.xlsx')
print("Injected NaN/Inf/None into test-supply.xlsx")
