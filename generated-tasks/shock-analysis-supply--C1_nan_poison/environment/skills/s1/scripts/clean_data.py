"""Clean NaN/Inf/None from XLSX using column-median fill, preserving sheet structure."""
import math
from openpyxl import load_workbook

path = '/root/test-supply.xlsx'
wb = load_workbook(path)

for ws in wb.worksheets:
    for col_idx in range(1, ws.max_column + 1):
        values = []
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if isinstance(cell.value, (int, float)) and math.isfinite(cell.value):
                values.append(cell.value)
        if not values:
            continue
        median = sorted(values)[len(values) // 2]
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value is None or (isinstance(cell.value, float) and (math.isnan(cell.value) or math.isinf(cell.value))):
                cell.value = median

wb.save(path)
print(f"Cleaned NaN/Inf in {path}")
