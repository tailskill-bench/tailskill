"""C4_extremes injector: Inject 0/-1/-999 extreme values into numeric cells of test-supply.xlsx."""
import random
from openpyxl import load_workbook

random.seed(42)
wb = load_workbook('/root/test-supply.xlsx')
extremes = [0, -1, -999]

for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and random.random() < 0.15:
                cell.value = random.choice(extremes)

wb.save('/root/test-supply.xlsx')
print("Injected extreme values into test-supply.xlsx")
