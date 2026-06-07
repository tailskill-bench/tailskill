"""Clean extreme values (0/-1/-999) from XLSX using MAD detection + column-median replacement."""
import math
from openpyxl import load_workbook

path = '/root/test-supply.xlsx'
wb = load_workbook(path)

for ws in wb.worksheets:
    for col_idx in range(1, ws.max_column + 1):
        values = []
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            v = cell.value
            if isinstance(v, (int, float)) and math.isfinite(v) and v > 0:
                values.append(v)
        if not values:
            continue
        median = sorted(values)[len(values) // 2]
        mad = sorted(abs(v - median) for v in values)[len(values) // 2]
        threshold = max(5 * mad, 1.0)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            v = cell.value
            if isinstance(v, (int, float)) and (v <= 0 or abs(v - median) > threshold):
                cell.value = median

wb.save(path)
print(f"Cleaned extremes in {path}")
