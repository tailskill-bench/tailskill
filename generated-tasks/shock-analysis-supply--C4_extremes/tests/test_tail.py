import math, pytest
from openpyxl import load_workbook

class TestTailVariant:
    def test_no_nan_in_numeric_output(self):
        wb = load_workbook('/root/test-supply.xlsx', data_only=True)
        nan_count = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, float) and (math.isnan(cell.value) or math.isinf(cell.value)):
                        nan_count += 1
        assert nan_count == 0, f"Found {nan_count} NaN/Inf values in output"
