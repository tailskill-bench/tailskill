"""C5_type_confusion injector: Convert numeric cells to string representations in test-supply.xlsx."""
import random
from openpyxl import load_workbook

random.seed(42)
wb = load_workbook('/root/test-supply.xlsx')

for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and random.random() < 0.2:
                cell.value = str(cell.value)

wb.save('/root/test-supply.xlsx')
print("Converted numeric cells to strings in test-supply.xlsx")
