"""Detect and convert string-represented numbers back to numeric types in XLSX."""
from openpyxl import load_workbook

path = '/root/test-supply.xlsx'
wb = load_workbook(path)

for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str):
                try:
                    v = float(cell.value)
                    cell.value = int(v) if v == int(v) else v
                except (ValueError, TypeError):
                    pass

wb.save(path)
print(f"Converted string numbers back to numeric in {path}")
