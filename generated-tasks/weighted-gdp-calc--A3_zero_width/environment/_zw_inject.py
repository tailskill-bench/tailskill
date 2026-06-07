#!/usr/bin/env python3
"""Inject zero-width characters into XLSX Data sheet numeric values."""
import random, sys
from openpyxl import load_workbook

random.seed(42)

filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/gdp.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 5  # 1-indexed column
ratio = 0.03
zw_chars = ["\u200b", "\u200c", "\u200d"]

wb = load_workbook(filepath)
ws = wb['Data']

rows = list(range(21, 41))
n_inject = max(1, int(len(rows) * ratio))
targets = random.sample(rows, min(n_inject, len(rows)))

count = 0
for row in targets:
    val = ws.cell(row=row, column=col_idx).value
    if val is not None and isinstance(val, (int, float)):
        s = str(val)
        mid = len(s) // 2
        ws.cell(row=row, column=col_idx).value = s[:mid] + random.choice(zw_chars) + s[mid:]
        count += 1

wb.save(filepath)
print(f"[TailSkills] Zero-width: {count} cells modified in Data sheet")
