#!/usr/bin/env python3
"""Strip zero-width Unicode characters from XLSX Data sheet cells."""
import unicodedata, sys
from openpyxl import load_workbook

filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/gdp.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 4  # 0-indexed for iter_rows

wb = load_workbook(filepath)
ws = wb['Data']

count = 0
for row in ws.iter_rows(min_row=21, max_row=40, values_only=False):
    cell = row[col_idx]
    val = cell.value
    if isinstance(val, str):
        cleaned = ''.join(c for c in val if unicodedata.category(c) not in ('Cf', 'Cc'))
        try:
            cell.value = float(cleaned) if '.' in cleaned else int(cleaned)
            count += 1
        except (ValueError, TypeError):
            cell.value = None

wb.save(filepath)
wb.close()
print(f"Cleaned {count} cells with zero-width characters")
