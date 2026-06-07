import random, math
random.seed(42)
from openpyxl import load_workbook

wb = load_workbook('/root/gdp.xlsx')
ws = wb['Data']
# Target rows 21-40 (numeric data), column E (1-indexed 5, 0-indexed 4)
col_idx = 5  # 1-indexed
data_rows = list(range(21, 41))
n_inject = max(1, int(len(data_rows) * 0.02))
targets = random.sample(data_rows, min(n_inject, len(data_rows)))
for row_idx in targets:
    cell = ws.cell(row=row_idx, column=col_idx)
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.value = float('nan')
wb.save('/root/gdp.xlsx')
print(f'[TailSkills] NaN: {len(targets)} cells modified in Data sheet column {col_idx}')
