#!/usr/bin/env python3
"""Replace NaN/Inf values with row-level median of other year columns."""
import sys, math
from openpyxl import load_workbook

target = sys.argv[1] if len(sys.argv) > 1 else "/root/gdp.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 4  # 0-indexed

wb = load_workbook(target)
ws = wb["Data"]

# Find year columns from row 4
year_cols = []
for c in range(1, 20):
    val = ws.cell(row=4, column=c).value
    if isinstance(val, (int, float)) and 2000 <= val <= 2100:
        year_cols.append(c)

target_col = col_idx + 1  # convert 0-indexed to 1-indexed
count = 0
for row_idx in range(21, 41):
    cell = ws.cell(row=row_idx, column=target_col)
    val = cell.value
    if val is None or (isinstance(val, float) and not math.isfinite(val)):
        other_vals = []
        for c in year_cols:
            if c == target_col:
                continue
            v = ws.cell(row=row_idx, column=c).value
            if v is not None and isinstance(v, (int, float)) and math.isfinite(v):
                other_vals.append(v)
        if other_vals:
            median = sorted(other_vals)[len(other_vals) // 2]
            cell.value = median
            count += 1

wb.save(target)
print(f"Replaced {count} NaN/Inf values with row-median")
