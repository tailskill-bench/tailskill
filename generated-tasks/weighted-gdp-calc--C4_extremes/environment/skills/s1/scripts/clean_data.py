#!/usr/bin/env python3
"""Replace extreme outlier values with row-level median using MAD + non-positive detection."""
import sys, math
from openpyxl import load_workbook

target = sys.argv[1] if len(sys.argv) > 1 else "/root/gdp.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 4  # 0-indexed

wb = load_workbook(target)
ws = wb["Data"]

# Find year columns from row 4
year_cols = []
for c in range(1, 20):
    val = ws.cell(row=4, column=c).value
    if isinstance(val, (int, float)) and 2000 <= val <= 2100:
        year_cols.append(c)

target_col = col_idx + 1  # 0-indexed to 1-indexed

# Compute column-level MAD for additional outlier detection
valid = []
for row_idx in range(21, 41):
    val = ws.cell(row=row_idx, column=target_col).value
    if val is not None and isinstance(val, (int, float)) and math.isfinite(val) and val > 0:
        valid.append(val)

if valid:
    median_col = sorted(valid)[len(valid) // 2]
    mad = sorted(abs(v - median_col) for v in valid)[len(valid) // 2]
    threshold = max(5 * mad, 1.0)
else:
    threshold = 1.0

count = 0
for row_idx in range(21, 41):
    cell = ws.cell(row=row_idx, column=target_col)
    val = cell.value
    corrupted = False
    if val is None:
        corrupted = True
    elif isinstance(val, str):
        corrupted = True
    elif isinstance(val, (int, float)):
        if not math.isfinite(val) or val <= 0 or abs(val - median_col) > threshold:
            corrupted = True

    if corrupted:
        other_vals = []
        for c in year_cols:
            if c == target_col:
                continue
            v = ws.cell(row=row_idx, column=c).value
            if v is not None and isinstance(v, (int, float)) and math.isfinite(v) and v > 0:
                other_vals.append(v)
        if other_vals:
            median = sorted(other_vals)[len(other_vals) // 2]
            cell.value = median
            count += 1

wb.save(target)
print(f"Replaced {count} extreme values with row-median (threshold={threshold:.2f})")
