import random
random.seed(42)
from openpyxl import load_workbook

wb = load_workbook('/root/gdp.xlsx')
ws = wb['Data']
col_idx = 5  # 1-indexed column E
markers = ["N/A", "n/a", "-", "TBD", "#REF!"]
data_rows = random.sample(range(21, 41), 5)
for i, row_idx in enumerate(data_rows):
    cell = ws.cell(row=row_idx, column=col_idx)
    if cell.value is not None:
        cell.value = markers[i % len(markers)]
wb.save('/root/gdp.xlsx')
print(f'[TailSkills] Type confusion: {len(data_rows)} cells modified in Data sheet column {col_idx}')
