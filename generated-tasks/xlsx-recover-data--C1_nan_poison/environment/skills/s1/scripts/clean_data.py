#!/usr/bin/env python3
"""Replace corrupted cells in Budget sheet with column-median (same directorate, other years)."""
import sys, math
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/nasa_budget_incomplete.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 4  # 0-indexed

wb = load_workbook(filepath)
target_col = col_idx + 1  # 1-indexed (column E = 5)

ws = wb["Budget by Directorate"] if "Budget by Directorate" in wb.sheetnames else wb.active

# Compute column median from valid numeric values in target column
valid = []
for r in range(1, ws.max_row + 1):
    cell = ws.cell(row=r, column=target_col)
    if isinstance(cell, MergedCell):
        continue
    v = cell.value
    if v is not None and isinstance(v, (int, float)) and math.isfinite(v):
        valid.append(v)

col_median = sorted(valid)[len(valid) // 2] if valid else 0

count = 0
for row_idx in range(1, ws.max_row + 1):
    cell = ws.cell(row=row_idx, column=target_col)
    if isinstance(cell, MergedCell):
        continue
    val = cell.value
    if val is None or (isinstance(val, float) and not math.isfinite(val)) or isinstance(val, str):
        cell.value = col_median
        count += 1

wb.save(filepath)
wb.close()
print(f"Replaced {count} corrupted values with column-median ({col_median})")
