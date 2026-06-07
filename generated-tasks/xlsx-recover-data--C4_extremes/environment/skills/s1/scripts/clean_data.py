#!/usr/bin/env python3
"""Replace extreme values in Budget sheet with column-median using MAD + non-positive detection."""
import sys, math
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/nasa_budget_incomplete.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 4  # 0-indexed

wb = load_workbook(filepath)
target_col = col_idx + 1  # 1-indexed

ws = wb["Budget by Directorate"] if "Budget by Directorate" in wb.sheetnames else wb.active

# Compute column-level MAD
valid = []
for r in range(1, ws.max_row + 1):
    cell = ws.cell(row=r, column=target_col)
    if isinstance(cell, MergedCell):
        continue
    v = cell.value
    if v is not None and isinstance(v, (int, float)) and math.isfinite(v) and v > 0:
        valid.append(v)

if valid:
    median_col = sorted(valid)[len(valid) // 2]
    mad = sorted(abs(v - median_col) for v in valid)[len(valid) // 2]
    threshold = max(5 * mad, 1.0)
else:
    median_col = 0
    threshold = 1.0

count = 0
for row_idx in range(1, ws.max_row + 1):
    cell = ws.cell(row=row_idx, column=target_col)
    if isinstance(cell, MergedCell):
        continue
    val = cell.value
    corrupted = False
    if val is None:
        corrupted = True
    elif isinstance(val, str):
        corrupted = True
    elif isinstance(val, (int, float)):
        if not math.isfinite(val) or val <= 0 or abs(val - median_col) > threshold:
            corrupted = True

    if corrupted:
        cell.value = median_col
        count += 1

wb.save(filepath)
wb.close()
print(f"Replaced {count} extreme values with column-median ({median_col}, threshold={threshold:.2f})")
