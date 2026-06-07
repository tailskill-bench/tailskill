#!/usr/bin/env python3
"""Inject type confusion markers into Budget sheet numeric cells (column E)."""
import random, sys
from openpyxl import load_workbook

random.seed(42)
filepath = sys.argv[1] if len(sys.argv) > 1 else "/root/nasa_budget_incomplete.xlsx"
col_idx = int(sys.argv[2]) if len(sys.argv) > 2 else 5
markers = ["N/A", "n/a", "-", "#REF!", "TBD"]

wb = load_workbook(filepath)
ws = wb["Budget by Directorate"]

numeric_rows = [r for r in range(1, ws.max_row + 1)
                if isinstance(ws.cell(row=r, column=col_idx).value, (int, float))]

targets = random.sample(numeric_rows, min(5, len(numeric_rows)))
for i, row in enumerate(targets):
    ws.cell(row=row, column=col_idx).value = markers[i % len(markers)]

wb.save(filepath)
print(f"[TailSkills] Type confusion: {len(targets)} cells modified in Budget sheet col {col_idx}")
