"""
test_augmentor.py — Adds tail-variant specific tests and patches test.sh.
"""

from pathlib import Path

# ─── Available tail test functions ─────────────────────────────

TAIL_TESTS = {
    "test_no_nan_in_numeric_output": '''
    def test_no_nan_in_numeric_output(self):
        """Output should not contain NaN values even if input has them."""
        import os
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not available")

        for f in os.listdir('/root'):
            if f.endswith('.xlsx') and not f.startswith('.'):
                wb = openpyxl.load_workbook(f'/root/{f}', data_only=True)
                for ws in wb.worksheets:
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                        for col_idx, cell in enumerate(row):
                            if isinstance(cell, float):
                                assert cell == cell, (
                                    f"NaN found in {f}, sheet '{ws.title}', "
                                    f"row {row_idx}, col {col_idx}"
                                )
''',

    "test_no_bom_in_output": '''
    def test_no_bom_in_output(self):
        """Output files should not contain BOM markers."""
        import os
        for f in os.listdir('/root'):
            if f.endswith(('.csv', '.txt', '.json')):
                with open(f'/root/{f}', 'rb') as fh:
                    head = fh.read(3)
                    assert head != b'\\xef\\xbb\\xbf', f"BOM found in output: {f}"
''',

    "test_output_encoding_clean": '''
    def test_output_encoding_clean(self):
        """Output files should be valid UTF-8 without NUL bytes."""
        import os
        for f in os.listdir('/root'):
            if f.endswith(('.csv', '.txt', '.json')):
                with open(f'/root/{f}', 'rb') as fh:
                    content = fh.read()
                    assert b'\\x00' not in content, f"NUL byte in {f}"
                    try:
                        content.decode('utf-8')
                    except UnicodeDecodeError:
                        pytest.fail(f"Output {f} is not valid UTF-8")
''',

    "test_output_file_exists": '''
    def test_output_file_exists(self):
        """At least one output file must exist."""
        import os
        output_files = [f for f in os.listdir('/root')
                       if f.endswith(('.xlsx', '.csv', '.json', '.pdf'))
                       and not f.startswith('.')]
        assert len(output_files) > 0, "No output files found in /root"
''',

    "test_no_inf_in_output": '''
    def test_no_inf_in_output(self):
        """Output should not contain Inf values."""
        import math, os
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not available")

        for f in os.listdir('/root'):
            if f.endswith('.xlsx') and not f.startswith('.'):
                wb = openpyxl.load_workbook(f'/root/{f}', data_only=True)
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, float):
                                assert math.isfinite(cell), (
                                    f"Inf found in {f}, sheet '{ws.title}'"
                                )
''',
}

# ─── Test file template ──────────────────────────────────────

TAIL_TEST_TEMPLATE = '''"""
TailSkills variant test — auto-generated.
Variant: {variant_id}
These tests verify the skill produces correct output under variant conditions.
"""

import os
import pytest


class TestTailVariant:
    """Tail-variant specific assertions."""
{test_methods}
'''

# ─── test.sh patch ───────────────────────────────────────────

TEST_SH_PATCH = """
# ═══ TailSkills: run variant-specific tests ═══
if [ -f /tests/test_tail.py ]; then
    echo "=== TailSkills variant tests ==="
    pytest /tests/test_tail.py -rA -v --tb=short 2>&1 | tee -a /logs/verifier/tail_results.txt
fi
"""


class TestAugmentor:
    """Adds tail-variant test files and patches test.sh."""

    def __init__(self, tests_dir: str | Path):
        self.tests_dir = Path(tests_dir)

    def add_tail_test(self, test_name: str, variant_id: str):
        """Add a tail test file with the specified test function."""
        if test_name not in TAIL_TESTS:
            print(f"[TailSkills] Warning: Unknown test '{test_name}', skipping")
            return

        test_method = TAIL_TESTS[test_name]
        content = TAIL_TEST_TEMPLATE.format(
            variant_id=variant_id,
            test_methods=test_method,
        )

        test_file = self.tests_dir / "test_tail.py"

        if test_file.exists():
            # Append the new test method to existing file
            with open(test_file, "r", encoding="utf-8") as f:
                existing = f.read()
            if test_name not in existing:
                # Add the method inside the existing class
                existing = existing.rstrip() + "\n" + test_method + "\n"
                with open(test_file, "w", encoding="utf-8") as f:
                    f.write(existing)
        else:
            with open(test_file, "w", encoding="utf-8") as f:
                f.write(content)

    def patch_test_sh(self):
        """Append tail test runner to test.sh."""
        test_sh = self.tests_dir / "test.sh"
        if not test_sh.exists():
            return

        with open(test_sh, "r", encoding="utf-8") as f:
            content = f.read()

        if "TailSkills" in content:
            return  # Already patched

        # Insert before the last 'exit 0' or at the end
        if "exit 0" in content:
            content = content.replace("exit 0", TEST_SH_PATCH.strip() + "\nexit 0", 1)
        else:
            content += "\n" + TEST_SH_PATCH

        with open(test_sh, "w", encoding="utf-8") as f:
            f.write(content)
